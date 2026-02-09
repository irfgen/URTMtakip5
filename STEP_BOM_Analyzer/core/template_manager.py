"""
STEP BOM Analyzer - Custom Template Manager
Handles user-defined BOM templates and report customization.
"""

import os
import json
import jinja2
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass, asdict
from enum import Enum
import logging
from datetime import datetime
import yaml
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class TemplateFormat(Enum):
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"
    YAML = "yaml"
    PDF = "pdf"


class TemplateType(Enum):
    BOM_REPORT = "bom_report"
    PART_GALLERY = "part_gallery"
    ASSEMBLY_TREE = "assembly_tree"
    MATERIAL_SUMMARY = "material_summary"
    COST_ANALYSIS = "cost_analysis"
    CUSTOM = "custom"


@dataclass
class TemplateConfig:
    """Template configuration structure"""
    name: str
    description: str
    template_type: TemplateType
    format: TemplateFormat
    template_content: str
    fields: List[str]
    filters: Dict[str, Any]
    styling: Dict[str, Any]
    metadata: Dict[str, Any]
    created_date: datetime
    modified_date: datetime
    author: str = "STEP BOM Analyzer"
    version: str = "1.0"


@dataclass
class BomData:
    """Standardized BOM data structure"""
    parts: List[Dict[str, Any]]
    assemblies: List[Dict[str, Any]]
    materials: List[Dict[str, Any]]
    metadata: Dict[str, Any]
    statistics: Dict[str, Any]
    hierarchy: Dict[str, Any]


class TemplateManager:
    """Advanced template management system"""
    
    def __init__(self, template_dir: Path = None):
        self.template_dir = template_dir or Path("templates")
        self.template_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        for template_type in TemplateType:
            (self.template_dir / template_type.value).mkdir(exist_ok=True)
        
        # Jinja2 environment
        self.jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(str(self.template_dir)),
            autoescape=jinja2.select_autoescape(['html', 'xml'])
        )
        
        # Add custom filters
        self._setup_custom_filters()
        
        # Template registry
        self.templates: Dict[str, TemplateConfig] = {}
        self._load_templates()
        
        # Default templates
        self._ensure_default_templates()
        
        self.logger = logging.getLogger(__name__)
    
    def _setup_custom_filters(self):
        """Setup custom Jinja2 filters"""
        
        def format_number(value, decimals=2):
            """Format number with specified decimals"""
            try:
                return f"{float(value):.{decimals}f}"
            except (ValueError, TypeError):
                return str(value)
        
        def format_size(bytes_value):
            """Format byte size in human readable format"""
            try:
                bytes_value = float(bytes_value)
                for unit in ['B', 'KB', 'MB', 'GB']:
                    if bytes_value < 1024.0:
                        return f"{bytes_value:.1f} {unit}"
                    bytes_value /= 1024.0
                return f"{bytes_value:.1f} TB"
            except (ValueError, TypeError):
                return str(bytes_value)
        
        def sort_by_field(items, field_name):
            """Sort items by field name"""
            try:
                return sorted(items, key=lambda x: x.get(field_name, ''))
            except (TypeError, AttributeError):
                return items
        
        def group_by_field(items, field_name):
            """Group items by field name"""
            groups = {}
            for item in items:
                key = item.get(field_name, 'Unknown')
                if key not in groups:
                    groups[key] = []
                groups[key].append(item)
            return groups
        
        def calculate_total(items, field_name):
            """Calculate total of numeric field"""
            try:
                return sum(float(item.get(field_name, 0)) for item in items)
            except (ValueError, TypeError):
                return 0
        
        self.jinja_env.filters['format_number'] = format_number
        self.jinja_env.filters['format_size'] = format_size
        self.jinja_env.filters['sort_by'] = sort_by_field
        self.jinja_env.filters['group_by'] = group_by_field
        self.jinja_env.filters['total'] = calculate_total
    
    def _load_templates(self):
        """Load existing templates from disk"""
        for template_type in TemplateType:
            type_dir = self.template_dir / template_type.value
            if type_dir.exists():
                for config_file in type_dir.glob("*.json"):
                    try:
                        with open(config_file, 'r', encoding='utf-8') as f:
                            config_data = json.load(f)
                        
                        # Convert dates
                        config_data['created_date'] = datetime.fromisoformat(
                            config_data['created_date']
                        )
                        config_data['modified_date'] = datetime.fromisoformat(
                            config_data['modified_date']
                        )
                        config_data['template_type'] = TemplateType(config_data['template_type'])
                        config_data['format'] = TemplateFormat(config_data['format'])
                        
                        template_config = TemplateConfig(**config_data)
                        self.templates[template_config.name] = template_config
                        
                    except Exception as e:
                        self.logger.error(f"Error loading template {config_file}: {e}")
    
    def _ensure_default_templates(self):
        """Create default templates if they don't exist"""
        default_templates = [
            self._create_default_bom_html_template(),
            self._create_default_part_gallery_template(),
            self._create_default_assembly_tree_template(),
            self._create_default_material_summary_template()
        ]
        
        for template_config in default_templates:
            if template_config.name not in self.templates:
                self.save_template(template_config)
    
    def _create_default_bom_html_template(self) -> TemplateConfig:
        """Create default HTML BOM template"""
        template_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ metadata.project_name }} - BOM Report</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 20px; }
        .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                 color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
                gap: 15px; margin: 20px 0; }
        .stat-card { background: #f8f9fa; padding: 15px; border-radius: 8px; 
                    border-left: 4px solid #007bff; }
        .stat-value { font-size: 24px; font-weight: bold; color: #007bff; }
        .stat-label { color: #6c757d; margin-top: 5px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #f8f9fa; font-weight: 600; }
        tr:hover { background-color: #f5f5f5; }
        .material-chip { display: inline-block; padding: 4px 8px; 
                        background-color: #e7f3ff; color: #0066cc; 
                        border-radius: 12px; font-size: 12px; margin: 2px; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ metadata.project_name | default('STEP BOM Analysis') }}</h1>
        <p>Generated: {{ metadata.generation_date }}</p>
        <p>File: {{ metadata.source_file }}</p>
    </div>

    <div class="stats">
        <div class="stat-card">
            <div class="stat-value">{{ statistics.total_parts }}</div>
            <div class="stat-label">Total Parts</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{{ statistics.unique_parts }}</div>
            <div class="stat-label">Unique Parts</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{{ statistics.total_assemblies }}</div>
            <div class="stat-label">Assemblies</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{{ materials | length }}</div>
            <div class="stat-label">Materials</div>
        </div>
    </div>

    <h2>Bill of Materials</h2>
    <table>
        <thead>
            <tr>
                <th>Part Name</th>
                <th>Quantity</th>
                <th>Material</th>
                <th>Volume (mm³)</th>
                <th>Mass (g)</th>
                <th>Assembly</th>
            </tr>
        </thead>
        <tbody>
            {% for part in parts | sort_by('name') %}
            <tr>
                <td><strong>{{ part.name }}</strong></td>
                <td>{{ part.quantity }}</td>
                <td>{{ part.material | default('Unknown') }}</td>
                <td>{{ part.volume | format_number(0) }}</td>
                <td>{{ part.mass | format_number(2) }}</td>
                <td>{{ part.assembly | default('Root') }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <h2>Material Summary</h2>
    {% for material_name, material_parts in parts | group_by('material') %}
    <h3>{{ material_name }}</h3>
    <ul>
        {% for part in material_parts %}
        <li>{{ part.name }} ({{ part.quantity }}x) - {{ part.mass | format_number(2) }}g</li>
        {% endfor %}
    </ul>
    <p><strong>Total Mass: {{ material_parts | total('mass') | format_number(2) }}g</strong></p>
    {% endfor %}
    
    <footer style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; 
                  color: #6c757d; font-size: 12px; text-align: center;">
        Generated by STEP BOM Analyzer v3.0 - {{ metadata.generation_date }}
    </footer>
</body>
</html>"""
        
        return TemplateConfig(
            name="default_bom_html",
            description="Default HTML BOM report template",
            template_type=TemplateType.BOM_REPORT,
            format=TemplateFormat.HTML,
            template_content=template_content,
            fields=["name", "quantity", "material", "volume", "mass", "assembly"],
            filters={"material": "all", "min_quantity": 0},
            styling={"theme": "modern", "colors": "blue"},
            metadata={"category": "standard", "tags": ["html", "responsive"]},
            created_date=datetime.now(),
            modified_date=datetime.now()
        )
    
    def _create_default_part_gallery_template(self) -> TemplateConfig:
        """Create default part gallery template"""
        template_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Part Gallery - {{ metadata.project_name }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .gallery { display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); 
                  gap: 20px; margin: 20px 0; }
        .part-card { background: white; border-radius: 8px; overflow: hidden; 
                    box-shadow: 0 2px 8px rgba(0,0,0,0.1); transition: transform 0.2s; }
        .part-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.15); }
        .part-image { width: 100%; height: 200px; object-fit: cover; background-color: #f0f0f0; }
        .part-info { padding: 15px; }
        .part-name { font-size: 18px; font-weight: bold; margin-bottom: 8px; }
        .part-details { font-size: 14px; color: #666; }
        .detail-row { display: flex; justify-content: space-between; margin: 4px 0; }
    </style>
</head>
<body>
    <h1>Part Gallery - {{ metadata.project_name }}</h1>
    
    <div class="gallery">
        {% for part in parts %}
        <div class="part-card">
            {% if part.image_path %}
            <img src="{{ part.image_path }}" alt="{{ part.name }}" class="part-image">
            {% else %}
            <div class="part-image" style="display: flex; align-items: center; justify-content: center; 
                                          color: #999;">No Image</div>
            {% endif %}
            
            <div class="part-info">
                <div class="part-name">{{ part.name }}</div>
                <div class="part-details">
                    <div class="detail-row">
                        <span>Quantity:</span>
                        <span>{{ part.quantity }}</span>
                    </div>
                    <div class="detail-row">
                        <span>Material:</span>
                        <span>{{ part.material | default('Unknown') }}</span>
                    </div>
                    <div class="detail-row">
                        <span>Volume:</span>
                        <span>{{ part.volume | format_number(0) }} mm³</span>
                    </div>
                    <div class="detail-row">
                        <span>Mass:</span>
                        <span>{{ part.mass | format_number(2) }} g</span>
                    </div>
                </div>
            </div>
        </div>
        {% endfor %}
    </div>
</body>
</html>"""
        
        return TemplateConfig(
            name="default_part_gallery",
            description="Default part gallery with images",
            template_type=TemplateType.PART_GALLERY,
            format=TemplateFormat.HTML,
            template_content=template_content,
            fields=["name", "image_path", "quantity", "material", "volume", "mass"],
            filters={"show_images": True},
            styling={"layout": "grid", "card_style": "modern"},
            metadata={"category": "gallery", "tags": ["visual", "responsive"]},
            created_date=datetime.now(),
            modified_date=datetime.now()
        )
    
    def _create_default_assembly_tree_template(self) -> TemplateConfig:
        """Create default assembly tree template"""
        template_content = """{
    "project": "{{ metadata.project_name }}",
    "generated": "{{ metadata.generation_date }}",
    "source_file": "{{ metadata.source_file }}",
    "assembly_tree": {{ hierarchy | tojson | safe }},
    "statistics": {
        "total_parts": {{ statistics.total_parts }},
        "total_assemblies": {{ statistics.total_assemblies }},
        "max_depth": {{ statistics.max_depth | default(0) }}
    },
    "parts_by_assembly": {
        {% for assembly_name, assembly_parts in parts | group_by('assembly') %}
        "{{ assembly_name }}": [
            {% for part in assembly_parts %}
            {
                "name": "{{ part.name }}",
                "quantity": {{ part.quantity }},
                "material": "{{ part.material | default('Unknown') }}",
                "volume": {{ part.volume | default(0) }},
                "mass": {{ part.mass | default(0) }}
            }{% if not loop.last %},{% endif %}
            {% endfor %}
        ]{% if not loop.last %},{% endif %}
        {% endfor %}
    }
}"""
        
        return TemplateConfig(
            name="default_assembly_tree",
            description="Default JSON assembly tree structure",
            template_type=TemplateType.ASSEMBLY_TREE,
            format=TemplateFormat.JSON,
            template_content=template_content,
            fields=["hierarchy", "assembly", "parts"],
            filters={"include_empty_assemblies": False},
            styling={},
            metadata={"category": "structure", "tags": ["json", "hierarchy"]},
            created_date=datetime.now(),
            modified_date=datetime.now()
        )
    
    def _create_default_material_summary_template(self) -> TemplateConfig:
        """Create default material summary template"""
        template_content = """Material,Part Count,Total Quantity,Total Volume (mm³),Total Mass (g),Average Density (g/cm³)
{% for material_name, material_parts in parts | group_by('material') -%}
{{ material_name }},{{ material_parts | length }},{{ material_parts | total('quantity') }},{{ material_parts | total('volume') | format_number(0) }},{{ material_parts | total('mass') | format_number(2) }},{{ (material_parts | total('mass') / (material_parts | total('volume') / 1000)) | format_number(3) }}
{% endfor -%}"""
        
        return TemplateConfig(
            name="default_material_summary",
            description="Default CSV material summary",
            template_type=TemplateType.MATERIAL_SUMMARY,
            format=TemplateFormat.CSV,
            template_content=template_content,
            fields=["material", "quantity", "volume", "mass"],
            filters={"group_by_material": True},
            styling={},
            metadata={"category": "analysis", "tags": ["csv", "materials"]},
            created_date=datetime.now(),
            modified_date=datetime.now()
        )
    
    def save_template(self, template_config: TemplateConfig):
        """Save template configuration to disk"""
        # Create type directory
        type_dir = self.template_dir / template_config.template_type.value
        type_dir.mkdir(exist_ok=True)
        
        # Save template content
        template_file = type_dir / f"{template_config.name}.{template_config.format.value}"
        with open(template_file, 'w', encoding='utf-8') as f:
            f.write(template_config.template_content)
        
        # Save configuration
        config_file = type_dir / f"{template_config.name}.json"
        config_data = asdict(template_config)
        
        # Convert dates to ISO format
        config_data['created_date'] = template_config.created_date.isoformat()
        config_data['modified_date'] = template_config.modified_date.isoformat()
        config_data['template_type'] = template_config.template_type.value
        config_data['format'] = template_config.format.value
        
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=2, ensure_ascii=False)
        
        # Update registry
        self.templates[template_config.name] = template_config
        
        self.logger.info(f"Template saved: {template_config.name}")
    
    def load_template(self, template_name: str) -> Optional[TemplateConfig]:
        """Load template by name"""
        return self.templates.get(template_name)
    
    def list_templates(self, template_type: TemplateType = None) -> List[TemplateConfig]:
        """List available templates"""
        if template_type:
            return [t for t in self.templates.values() if t.template_type == template_type]
        return list(self.templates.values())
    
    def delete_template(self, template_name: str) -> bool:
        """Delete template"""
        template_config = self.templates.get(template_name)
        if not template_config:
            return False
        
        try:
            # Delete files
            type_dir = self.template_dir / template_config.template_type.value
            template_file = type_dir / f"{template_name}.{template_config.format.value}"
            config_file = type_dir / f"{template_name}.json"
            
            if template_file.exists():
                template_file.unlink()
            if config_file.exists():
                config_file.unlink()
            
            # Remove from registry
            del self.templates[template_name]
            
            self.logger.info(f"Template deleted: {template_name}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error deleting template {template_name}: {e}")
            return False
    
    def render_template(
        self, 
        template_name: str, 
        bom_data: BomData, 
        output_file: Path = None
    ) -> str:
        """Render template with BOM data"""
        template_config = self.templates.get(template_name)
        if not template_config:
            raise ValueError(f"Template not found: {template_name}")
        
        # Prepare template data
        template_data = {
            'parts': bom_data.parts,
            'assemblies': bom_data.assemblies,
            'materials': bom_data.materials,
            'metadata': bom_data.metadata,
            'statistics': bom_data.statistics,
            'hierarchy': bom_data.hierarchy
        }
        
        try:
            # Render with Jinja2
            template = self.jinja_env.from_string(template_config.template_content)
            rendered_content = template.render(**template_data)
            
            # Apply post-processing based on format
            if template_config.format == TemplateFormat.JSON:
                # Validate JSON
                json.loads(rendered_content)
            elif template_config.format == TemplateFormat.EXCEL:
                # Convert to Excel format
                rendered_content = self._render_excel_template(template_config, bom_data, output_file)
            
            # Save to file if requested
            if output_file:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(rendered_content)
                self.logger.info(f"Template rendered to: {output_file}")
            
            return rendered_content
            
        except Exception as e:
            self.logger.error(f"Error rendering template {template_name}: {e}")
            raise
    
    def _render_excel_template(
        self, 
        template_config: TemplateConfig, 
        bom_data: BomData, 
        output_file: Path
    ) -> str:
        """Render Excel template"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # BOM Sheet
        bom_sheet = wb.create_sheet("BOM")
        df = pd.DataFrame(bom_data.parts)
        
        # Add headers
        for r in dataframe_to_rows(df, index=False, header=True):
            bom_sheet.append(r)
        
        # Style headers
        for cell in bom_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Material Summary Sheet
        materials_sheet = wb.create_sheet("Materials")
        material_summary = []
        
        for material_name in set(part.get('material', 'Unknown') for part in bom_data.parts):
            material_parts = [p for p in bom_data.parts if p.get('material') == material_name]
            total_quantity = sum(p.get('quantity', 0) for p in material_parts)
            total_mass = sum(p.get('mass', 0) for p in material_parts)
            
            material_summary.append({
                'Material': material_name,
                'Parts': len(material_parts),
                'Total Quantity': total_quantity,
                'Total Mass (g)': total_mass
            })
        
        df_materials = pd.DataFrame(material_summary)
        for r in dataframe_to_rows(df_materials, index=False, header=True):
            materials_sheet.append(r)
        
        # Style materials sheet
        for cell in materials_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Save Excel file
        if output_file:
            wb.save(output_file)
        
        return str(output_file) if output_file else "Excel content generated"
    
    def create_custom_template(
        self,
        name: str,
        description: str,
        template_type: TemplateType,
        format: TemplateFormat,
        template_content: str,
        fields: List[str],
        filters: Dict[str, Any] = None,
        styling: Dict[str, Any] = None
    ) -> TemplateConfig:
        """Create new custom template"""
        template_config = TemplateConfig(
            name=name,
            description=description,
            template_type=template_type,
            format=format,
            template_content=template_content,
            fields=fields,
            filters=filters or {},
            styling=styling or {},
            metadata={"category": "custom", "tags": ["user-created"]},
            created_date=datetime.now(),
            modified_date=datetime.now()
        )
        
        self.save_template(template_config)
        return template_config


def template_manager_example():
    """Example usage of template manager"""
    
    # Initialize template manager
    manager = TemplateManager()
    
    # Example BOM data
    bom_data = BomData(
        parts=[
            {
                'name': 'Shaft',
                'quantity': 2,
                'material': 'Steel',
                'volume': 15000,
                'mass': 117.75,
                'assembly': 'Main Assembly'
            },
            {
                'name': 'Bearing',
                'quantity': 4,
                'material': 'Steel',
                'volume': 5000,
                'mass': 39.25,
                'assembly': 'Main Assembly'
            }
        ],
        assemblies=[{'name': 'Main Assembly', 'parts': 6}],
        materials=[{'name': 'Steel', 'density': 7.85}],
        metadata={
            'project_name': 'Test Project',
            'source_file': 'test.step',
            'generation_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        statistics={
            'total_parts': 6,
            'unique_parts': 2,
            'total_assemblies': 1
        },
        hierarchy={'Main Assembly': {'Shaft': 2, 'Bearing': 4}}
    )
    
    # Render default HTML template
    output = manager.render_template('default_bom_html', bom_data)
    print(f"HTML template rendered: {len(output)} characters")
    
    # Create custom template
    custom_template = manager.create_custom_template(
        name="simple_list",
        description="Simple part list",
        template_type=TemplateType.BOM_REPORT,
        format=TemplateFormat.HTML,
        template_content="""
        <h1>{{ metadata.project_name }}</h1>
        <ul>
        {% for part in parts %}
        <li>{{ part.name }} - {{ part.quantity }}x</li>
        {% endfor %}
        </ul>
        """,
        fields=["name", "quantity"]
    )
    
    # List all templates
    templates = manager.list_templates()
    print(f"Available templates: {[t.name for t in templates]}")


if __name__ == "__main__":
    template_manager_example()